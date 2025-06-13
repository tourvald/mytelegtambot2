
import time

import pandas as pd
import logging
import openpyxl
import random
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime
from archive import archive, load_archive, get_last_date
from bs4 import BeautifulSoup
from my_libs.libs_selenium import create_chrome_driver_object

from archive import get_key_link
from my_libs.libs_google_sheets import get_myphones_spreadsheet, get_mysells_spreadsheet

from selenium.webdriver.common.by import By

from mylibs import av_price_sdt, av_price_auto


def avito_parce(url):
    print (url)
    driver = create_chrome_driver_object()
    driver.get(url)
    time.sleep(3)
    for i in range(5):
        try:
            contents = driver.page_source
            soup = BeautifulSoup(contents, 'lxml')
            search_request = soup.find('input', {'data-marker': 'search-form/suggest'}).get('value')  # Название запроса
            break
        except Exception as e:
            driver.refresh()
            print ('Ждем 5 сек')
            time.sleep(5)
            print (e)
    price_list = []
    div_catalog_serp = soup.find('div', {'data-marker': 'catalog-serp'})  # Выделяем блок в котором хранятся все цены
    for price in div_catalog_serp.find_all('meta', {'itemprop': 'price'}):  # Перебираем подблоки блока div_catalog_serp
        price_list.append(price.get('content'))  # Достаем из них цены и добавляем в список цен
    av_price_std = av_price_sdt(price_list)
    av_price = f'{av_price_std}'
    print(av_price_std)
    archive(datetime.date.today(), url, av_price, search_request.lower())
    return av_price_std, search_request.lower()

def get_soup_for_avito_parce_old(url):
    with open('data/checked_proxies.txt', 'r', encoding='UTF-8') as f:
        proxies = f.readlines()
    print (proxies)
    proxies.reverse()
    proxy_cycle = 0
    retryes = 0
    len_proxies = len(proxies)
    driver = create_chrome_driver_object(proxy=proxies[proxy_cycle])
    while proxy_cycle != 2:
        retryes += 1
        try:
            driver.get(url)
            driver.implicitly_wait(10)
        except Exception as e:
            print(e.__cause__)
        soup = BeautifulSoup(driver.page_source, 'lxml')
        if not soup.title:
            print(f'Cуп не получен, попытка {retryes}, ждем  {delay * retryes} сек')
            if retryes == 7:
                retryes = 0
                driver.close()
                count_proxies += 1
                if count_proxies > len_proxies:
                    count_proxies = 0
                    proxy_cycle += 1
                    print(f'Загрузка не удалась. Попыток - {retryes}')
                driver = create_chrome_driver_object(proxy=proxies[count_proxies])
            continue

        if soup.title.text == 'Доступ временно заблокирован':
            print(soup.title.text)
            driver.close()
            count_proxies += 1
            if count_proxies > len_proxies:
                count_proxies = 0
                proxy_cycle += 1
                print(f'Загрузка не удалась. Попыток - {retryes}')
            driver = create_chrome_driver_object(proxy=proxies[count_proxies])
            delay = random.uniform(3, 6)
            time.sleep(delay)
            continue
        else:
            break
    return soup

def get_soup_for_avito_parce(url, driver, attempts=5):
    """
    Загружает страницу с использованием Selenium и возвращает объект BeautifulSoup.
    """
    delay = 5  # Начальная задержка
    for attempt in range(attempts):
        try:
            # Открываем страницу
            driver.get(url)

            # Рандомная задержка (эмуляция человеческого поведения)
            time.sleep(random.uniform(delay, delay + 2))

            # Проверка на блокировку по IP или загрузку страницы
            h2_elements = driver.find_elements(By.TAG_NAME, "h2")
            h2_texts = [element.text for element in h2_elements]

            if "Доступ ограничен: проблема с IP" in h2_texts or "Подождите, идет загрузка" in h2_texts:
                print(
                    f"Попытка {attempt + 1} для {url}: Обнаружена блокировка или проблема с загрузкой. Пробуем снова через {delay} секунд...")
                delay += 2  # Увеличиваем задержку
                continue  # Пробуем еще раз

            # Получаем HTML страницы
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            return soup

        except Exception as e:
            print(f"Произошла ошибка на попытке {attempt + 1} для {url}: {e}")
            delay += 2  # Увеличиваем задержку
            continue  # Пробуем еще раз

    # Если после всех попыток не удалось получить содержимое
    print(f"Не удалось загрузить страницу {url} корректно после всех попыток.")
    return None

def avito_parce_soup(soup):
    """
    Функция для парсинга страниц Avito и извлечения цен.

    :param soup: Объект BeautifulSoup, содержащий HTML-структуру страницы.
    :return: Кортеж со средней ценой (или сообщением об ошибке) и запросом поиска.
    """
    price_list = []

    # Получение названия запроса поиска из атрибута value
    search_input = soup.find('input', {'data-marker': 'search-form/suggest/input'})
    if search_input:
        search_request = search_input.get('value')
    else:
        search_request = "Не найдено название запроса"

    # Поиск блока, содержащего все цены
    div_catalog_serp = soup.find('div', {'data-marker': 'catalog-serp'})
    if div_catalog_serp:
        # Перебор всех подблоков и извлечение цен
        for price in div_catalog_serp.find_all('meta', {'itemprop': 'price'}):
            price_list.append(price.get('content'))  # Добавляем цену в список
    else:
        print("Не удалось найти блок с данными каталога.")

    # Расчет средней цены, если цены найдены
    if len(price_list) > 1:
        try:
            av_price_std = av_price_sdt(price_list)  # Функция для вычисления среднего и std
        except Exception as e:
            print(f"Ошибка при вычислении средней цены: {e}")
            av_price_std = "Ошибка вычисления"
    else:
        print('Длина списка цен меньше 1')
        av_price_std = "Цен не обнаружено"

    print(av_price_std)
    return av_price_std, search_request.lower()


def avito_auto_parce_soup(soup):
    price_list = []
    # Обновлённый селектор для поиска запроса
    search_request = soup.find('input', {'data-marker': 'search-form/suggest'}).get('value')

    # Обновлённый селектор для поиска блока каталога
    div_catalog_serp = soup.select_one('div.index-content-c0K1j > div.index-root-gtkvj')

    # Проверяем, нашли ли мы блок каталога
    if div_catalog_serp:
        for price in div_catalog_serp.find_all('meta', {'itemprop': 'price'}):
            price_list.append(price.get('content'))  # Достаем из них цены и добавляем в список цен
    else:
        print("Не удалось найти блок каталога.")

    # Средняя цена и стандартное отклонение
    av_price_std = av_price_auto(price_list)
    print(av_price_std)

    return av_price_std, search_request.lower()


def myphones_get_avarage_prices(range_="myphones"):
    start_time = time.perf_counter()
    sum_av_price = 0
    sum_sell_price = 0
    return_ = []

    # Создание драйвера в контекстном менеджере
    with create_chrome_driver_object(headless=True, proxy=False) as driver:
        myphones = get_myphones_spreadsheet(range=range_)

        for myphone in myphones['values']:
            key = myphone[1]
            key_link = myphone[3]
            index = myphone[2]

            # Получение данных с Avito
            soup = get_soup_for_avito_parce(key_link, driver, attempts=5)
            av_price, key = avito_parce_soup(soup)

            # Расчет цены продажи
            sellprice = int(av_price * float(index))

            # Формирование результата
            return_.append(f'{myphone[0]}, - {av_price}"/"{sellprice}')
            sum_av_price += av_price
            sum_sell_price += sellprice

        # Добавление общей суммы цен
        return_.append(f'Средняя цена - {sum_av_price}')
        return_.append(f'Примерная цена продажи - {sum_sell_price}')

    finish_time = time.perf_counter()
    print(f'Finished in {round(finish_time - start_time, 2)} second(s)')

    return return_
def mycars_get_avarage_prices():
    sum_av_price = 0
    sum_sell_price = 0
    return_ = {}
    # Создание драйвера в контекстном менеджере
    with create_chrome_driver_object(headless=True, proxy=False) as driver:
        myphones = get_myphones_spreadsheet(range='mycars')

        myphones = get_myphones_spreadsheet(range='mycars')
        for myphone in myphones['values']:
            key = myphone[1]
            key_link = myphone[3]
            index = myphone[2]
            soup = get_soup_for_avito_parce(key_link, driver, attempts=5)
            av_price, key = avito_parce_soup(soup)

            sellprice = int(int(av_price) * float(myphone[2]))
            return_[myphone[0]]= av_price
            sum_av_price += av_price
            sum_sell_price += sellprice
        return_['total'] = sum_av_price
        print(return_)
        return return_

# def mycars_get_avarage_prices_2():
#     sum_av_price = 0
#     sum_sell_price = 0
#     return_ = {}
#     # Создание драйвера в контекстном менеджере
#     with create_chrome_driver_object(headless=True, proxy=False) as driver:
#         myphones = get_myphones_spreadsheet(range='mycars')
#
#         myphones = get_myphones_spreadsheet(range='mycars')
#         for myphone in myphones['values']:
#             key = myphone[1]
#             key_link = myphone[3]
#             index = myphone[2]
#             soup = get_soup_for_avito_parce(key_link, driver, attempts=5)
#             av_price, key = avito_parce_soup(soup)
#
#             sellprice = int(int(av_price) * float(myphone[2]))
#             return_[myphone[0]]= av_price
#             sum_av_price += av_price
#             sum_sell_price += sellprice
#         return_['total'] = sum_av_price
#         print(return_)
#         return return_


logging.basicConfig(level=logging.INFO)


def get_keys_to_update(archive, days_threshold=14):
    """Получает список ключей, которые нужно обновить в архиве."""
    keys_to_update = []
    for key in archive:
        date = get_last_date(key)
        timedelta = (datetime.datetime.today() - datetime.datetime.strptime(date, '%Y-%m-%d')).days
        if timedelta > days_threshold:
            logging.info(f'Key {key} needs update, last update was {timedelta} days ago.')
            keys_to_update.append(key)
    return keys_to_update


def update_keys(driver, keys):
    """Обновляет данные для каждого ключа из списка."""
    updated_keys = []  # Список обновленных ключей
    for key in keys:
        try:
            url = get_key_link(key)
            logging.info(f'Updating key {key} with URL {url}')

            # Парсинг данных с использованием драйвера
            soup = get_soup_for_avito_parce(url, driver)
            av_price_std, search_request = avito_parce_soup(soup)

            # Обновление архива
            archive(datetime.date.today(), url, av_price_std, key)
            updated_keys.append(key)  # Добавление обновленного ключа в список

            # Задержка между запросами
            wait_between_requests()

        except Exception as e:
            logging.error(f'Error updating key {key}: {e}')
            continue

    return updated_keys  # Возвращаем список обновленных ключей


def wait_between_requests(min_delay=1, max_delay=10):
    """Ожидание между запросами для избегания блокировки IP."""
    delay = random.uniform(min_delay, max_delay)
    logging.info(f'Waiting for {delay:.2f} seconds before the next request')
    time.sleep(delay)


def update_archive(amount_of_keys: int):
    """Основная функция для обновления архива."""
    arch = load_archive()  # Загрузка архива
    total_keys = len(arch)  # Всего позиций в архиве

    # Получение списка ключей, которые нужно обновить
    keys_to_update = get_keys_to_update(arch)
    keys_to_update_count = len(keys_to_update)  # Сколько позиций требуют обновления

    # Ограничиваем количество ключей, если указано
    keys_to_update = keys_to_update[:amount_of_keys]

    # Инициализация драйвера в контекстном менеджере
    with create_chrome_driver_object() as driver:
        # Обновление данных для ключей
        updated_keys = update_keys(driver, keys_to_update)

    # Информация о процессе обновления
    logging.info('Archive update completed.')
    logging.info(f'Total keys in archive: {total_keys}')
    logging.info(f'Keys that require update: {keys_to_update_count}')
    logging.info(f'Keys updated in this run: {len(updated_keys)}')
    logging.info(f'Keys that were updated: {updated_keys}')
    logging.info(f'Keys still require update: {keys_to_update_count - len(updated_keys)}')


import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def write_car_data():
    # Получаем данные и обновляем дату и время
    outputs = mycars_get_avarage_prices()
    current_datetime = datetime.datetime.now()
    outputs['date'] = current_datetime.strftime("%d-%m-%Y")
    outputs['time'] = current_datetime.strftime("%H:%M")

    try:
        # Чтение данных из Excel
        df = pd.read_excel('data/mycars/mycars2.xlsx', engine='openpyxl')

        # Добавляем новые данные строкой сверху
        new_row = pd.DataFrame([outputs])
        df = pd.concat([new_row, df], ignore_index=True)

        # Устанавливаем порядок первых трех столбцов
        fixed_columns = ['date', 'time', 'total']

        # Все столбцы кроме первых трех
        other_columns = [col for col in df.columns if col not in fixed_columns]

        # Столбцы, присутствующие в новой выдаче
        present_columns = [col for col in other_columns if col in outputs]

        # Столбцы, отсутствующие в новой выдаче (их нужно переместить в конец)
        absent_columns = [col for col in other_columns if col not in outputs]

        # Новый порядок столбцов: сначала фиксированные, затем новые, затем удаленные
        new_column_order = fixed_columns + present_columns + absent_columns
        df = df[new_column_order]

        # Преобразуем тип данных для числовых столбцов
        for col in present_columns:
            df[col] = df[col].fillna(0).astype(int)

        # Удаляем столбцы, где первые 1000 значений равны нулю или пустые
        df = remove_columns_with_first_1000_zeros_or_empty(df)

        # Сохранение данных в Excel
        output_file_path = 'data/mycars/mycars2.xlsx'
        df.to_excel(output_file_path, engine='openpyxl', index=False)
        print("Данные успешно обновлены!")

        # Настройка ширины столбцов на основе содержимого
        adjust_column_width_based_on_content(output_file_path)

        # Удаление строк ниже 6000-й
        remove_rows_below_6000(output_file_path)

    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")


def remove_columns_with_first_1000_zeros_or_empty(df):
    # Определение столбцов для удаления
    columns_to_remove = [col for col in df.columns if df[col].head(1000).replace(0, pd.NA).isna().all()]

    # Удаление столбцов
    if columns_to_remove:
        df = df.drop(columns=columns_to_remove)
        print(f"Удалены столбцы: {', '.join(columns_to_remove)}")
    else:
        print("Нет столбцов для удаления.")

    return df


def adjust_column_width_based_on_content(file_path):
    # Загружаем существующую книгу
    wb = load_workbook(file_path)
    ws = wb.active  # Предполагается, что активный лист — тот, который нужно отредактировать

    # Устанавливаем ширину столбцов в соответствии с длиной самого длинного содержимого в каждом столбце
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Номер столбца (A, B, C...)
        column_letter = get_column_letter(column)  # Письменное представление номера столбца (A, B, C...)

        max_cell_value = ""  # Инициализируем переменную для хранения содержимого ячейки с максимальной длиной

        for cell in col:
            try:
                # Определяем длину содержимого ячейки и обновляем max_length и max_cell_value, если текущая длина больше
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                    max_cell_value = str(cell.value)
            except:
                pass

        # Определяем ширину столбца, прибавляя небольшое смещение для более удобного отображения
        adjusted_width = max(max_length + 2, 8)  # Минимальная ширина 8 для читабельности
        ws.column_dimensions[column_letter].width = adjusted_width

        # Выводим ширину столбца и информацию о ячейке с максимальной длиной на экран
        print(f"Столбец {column_letter} ({ws[column_letter + '1'].value}): ширина = {adjusted_width}")
        print(f"   Ячейка с максимальным количеством символов: '{max_cell_value}' ({max_length} символов)")

    # Сохраняем изменения
    wb.save(file_path)


def remove_rows_below_6000(file_path):
    # Загружаем существующую книгу
    wb = load_workbook(file_path)
    ws = wb.active  # Предполагается, что активный лист — тот, который нужно отредактировать

    # Определяем количество строк в таблице
    max_row = ws.max_row

    # Проверяем, есть ли строки ниже 6000-й
    if max_row > 6000:
        # Удаляем строки ниже 6000-й
        ws.delete_rows(6001, max_row - 6000)
        print(f"Удалены строки с 6001 по {max_row}.")
    else:
        print("В таблице менее 6000 строк, ничего не удалено.")

    # Сохраняем изменения
    wb.save(file_path)
    print("Изменения сохранены.")

def daily_mean():
    df = pd.read_csv('data/mycars/mycars.csv', encoding='utf-8', delimiter=',')
    return_ = date_mean = df.groupby('date').mean()['total'].astype(int)
    date_mean.to_frame().to_csv('data/mycars/mycarsreport.csv', encoding='utf-8', sep=',')
    date_mean.to_frame().to_csv('data/mycars/mycarsreport.csv', encoding='utf-8', sep=',', delimiter=';')
    return return_

def convert_old_db():
    df = pd.read_csv('data/mycars/mycars.csv', encoding='utf-8', sep=';')
    print(df.head())
    # df = df.sort_index(ascending=True)
    # df.to_csv(f'data/mycars/mycars2.csv', encoding='utf-8', sep=';', index=False)
    # print(df.head())

if __name__ == "__main__":
    import os;


    write_car_data()

